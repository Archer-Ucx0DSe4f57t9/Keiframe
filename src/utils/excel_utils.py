import pandas as pd
import datetime
import sys
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
from src.utils.temp_translate_utils import MUTATOR_HEADER_MAP, MAP_HEADER_MAP

class ExcelUtil:

    @staticmethod
    def _get_column_width(value):
        """计算单元格内容的显示宽度"""
        if value is None: return 0
        s = str(value)
        width = 0
        for char in s:
            width += 2 if ord(char) > 127 else 1.1
        return width

    @staticmethod
    def format_seconds_to_label(total_seconds):
        """格式化秒数为 mm:ss"""
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        return f"{minutes:02d}:{seconds:02d}"

    @staticmethod
    def export_configs(data_list, file_path, config_type='map'):
        """导出配置，强制 B 列为文本格式并优化列宽"""
        header_map = MAP_HEADER_MAP if config_type == 'map' else MUTATOR_HEADER_MAP
        df = pd.DataFrame(data_list)
        
        # 补齐标题列
        for col in header_map.keys():
            if col not in df.columns:
                df[col] = "" 
        
        df = df[list(header_map.keys())].rename(columns=header_map)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # --- 重点：固定 B 列（时间点）为文本格式 ---
        # 即使是空行也应用此格式，确保用户后续输入时不会被自动转换
        for row_idx in range(1, 1000): # 预设处理 1000 行
            ws.cell(row=row_idx, column=2).number_format = '@'

        # 自动调整列宽并解锁
        for i, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in col_cells:
                max_length = max(max_length, ExcelUtil._get_column_width(cell.value))
                cell.protection = Protection(locked=False)
            ws.column_dimensions[column_letter].width = max_length + 2

        ws.protection.sheet = True # 开启保护以维持格式，由于单元格 unlocked，用户仍可编辑
        wb.save(file_path)

    @staticmethod
    def import_configs(file_path, config_type='map'):
        """读取 Excel，并根据业务逻辑清洗数据"""
        header_map = MAP_HEADER_MAP if config_type == 'map' else MUTATOR_HEADER_MAP
        rev_map = {v: k for k, v in header_map.items()}
        
        identity_key = 'map_name' if config_type == 'map' else 'mutator_name'
        identity_chs = header_map[identity_key]

        try:
            # 强制将所有列读为 object，防止初次解析时丢失精度
            df = pd.read_excel(file_path, dtype=object)
            df = df.where(pd.notnull(df), None)

            if identity_chs not in df.columns:
                return None, f"❌ 找不到 '{identity_chs}' 列，请检查表头。"

            df = df.rename(columns=rev_map)
            data = df.to_dict(orient='records')

            results = []
            for row in data:
                # 过滤完全空的行
                if pd.isna(row.get(identity_key)): continue

                # 业务逻辑：非“净网行动”清空 count_value
                if str(row.get('map_name')).strip() != '净网行动':
                    row['count_value'] = None

                results.append(row)
            return results, None
        except Exception as e:
            return None, str(e)

    @staticmethod
    def parse_time_label(label):
        """智能解析时间 (支持 0.35 浮点数还原)"""
        if pd.isna(label): return 0
        try:
            parts = list(map(int, str(label).strip().split(':')))
            return parts[0] * 60 + parts[1] if len(parts) == 2 else (parts[0] * 3600 + parts[1] * 60 + parts[2] if len(parts) == 3 else int(float(label)))
        except: return 0
        
        
# ==========================================
# 2. 测试代码：显示全部每行所有标题列内容
# ==========================================
if __name__ == '__main__':
    args = sys.argv[1:]
    if not args:
        print("用法: python excel_utils.py [map|mutator] [文件路径(可选)]")
        sys.exit(0)

    cfg_type = args[0].lower()
    file_path = args[1] if len(args) > 1 else None

    if file_path:
        print(f"--- 正在执行 {cfg_type} 导入测试 (包含非净网行动清空逻辑) ---")
        data, err = ExcelUtil.import_configs(file_path, cfg_type)
        
        if err:
            print(err)
        else:
            header_map = MAP_HEADER_MAP if cfg_type == 'map' else MUTATOR_HEADER_MAP
            print(f"\n{'='*50}\n成功读取到 {len(data)} 条数据：\n{'='*50}")

            for i, row in enumerate(data):
                print(f"\n[第 {i+1} 行内容]")
                for key, chs_name in header_map.items():
                    val = row.get(key, "")
                    if pd.isna(val) or val is None: val = ""
                    print(f"  {chs_name}: {val}")
                
                # 特别核对系统生成值
                if 'time_value' in row:
                    print(f"  (系统解析 time_value: {row['time_value']} 秒)")
            
            print(f"\n{'='*50}\n测试结束。")

    else:
        # 导出 Mock 包含净网和其他地图，方便测试清空逻辑
        mock_data = {
            'map': [
                {'map_name': '净网行动', 'time_label': '01:20', 'count_value': 1, 'event_text': '节点1开始'},
                {'map_name': '克哈裂痕', 'time_label': '05:30', 'count_value': 99, 'event_text': '此行count_value导入后应被清空'}
            ],
            'mutator': [{'mutator_name': 'SpeedFreaks', 'time_label': '00:00', 'content_text': '速度狂魔'}]
        }
        target_file = f"test_cleanup_{cfg_type}.xlsx"
        ExcelUtil.export_configs(mock_data.get(cfg_type, []), target_file, cfg_type)
        print(f"已生成包含混合数据的模版: {target_file}")
        print(f"请运行: python excel_utils.py {cfg_type} {target_file} 观察'克哈裂痕'的计数值是否被清空。")